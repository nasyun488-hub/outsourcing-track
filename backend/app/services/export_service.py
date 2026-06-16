from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from datetime import datetime, timedelta
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.order import Order, OrderStatus
from app.models.process import Process
from app.models.record import ProcessRecord
from app.models.factory import Factory


class ExportService:
    """Excel导出服务"""

    # 样式定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    OVERDUE_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    NORMAL_FONT = Font(name="微软雅黑", size=10)
    HEADER_FONT_STYLE = Font(name="微软雅黑", size=10, bold=True)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, db: Session):
        self.db = db

    def export_to_excel(
        self,
        user_role: str,
        user_factory_id: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        factory_id: Optional[str] = None,
        status: Optional[str] = None,
        order_id: Optional[str] = None,
    ) -> io.BytesIO:
        """
        导出Excel
        - 权限过滤：企业管理员全量，厂家仅本厂
        """
        # 构建查询
        query = self.db.query(Order)

        # 权限过滤：企业管理员全量；其他角色只能导出本厂作为主厂或承接厂参与的订单
        if user_role != "enterprise_admin":
            order_ids = self.db.query(ProcessRecord.order_id).filter(ProcessRecord.factory_id == user_factory_id)
            query = query.filter((Order.primary_factory_id == user_factory_id) | (Order.order_id.in_(order_ids)))
        elif factory_id:
            query = query.filter(Order.primary_factory_id == factory_id)

        # 日期过滤
        if start_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Order.created_at >= start_dt)
        if end_date:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            query = query.filter(Order.created_at <= end_dt)

        # 状态过滤
        if status:
            query = query.filter(Order.order_status == status)
        if order_id:
            query = query.filter(Order.order_id.like(f"%{order_id}%"))

        orders = query.order_by(Order.updated_at.desc()).all()

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "订单总览"

        # 写入表头
        headers = [
            "订单ID", "主厂家ID", "订单状态", "总数量",
            "MOM创建时间", "创建时间", "更新时间",
            "是否有超期工序"
        ]
        self._write_header(ws, headers)

        # 写入数据
        overdue_threshold = timedelta(hours=48)
        now = datetime.now()

        for row_idx, order in enumerate(orders, start=2):
            has_overdue = self._check_order_has_overdue(order.order_id, now, overdue_threshold)

            row_data = [
                order.order_id,
                order.primary_factory_id,
                self._get_status_text(order.order_status),
                order.total_qty,
                order.mom_created_at.strftime("%Y-%m-%d %H:%M:%S") if order.mom_created_at else "",
                order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                order.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
                "是" if has_overdue else "否",
            ]
            self._write_row(ws, row_idx, row_data, is_overdue=has_overdue)

        # 创建工序详情sheet
        ws2 = wb.create_sheet(title="工序明细")
        self._write_processes_sheet(ws2, orders, now, overdue_threshold)

        # 调整列宽
        self._adjust_column_width(ws, headers)
        self._adjust_column_width(ws2, [
            "订单ID", "工序ID", "工序名称", "序号", "状态",
            "接收时间", "发出时间", "是否超期"
        ])

        # 保存到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_header(self, ws, headers: List[str]):
        """写入表头"""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER

    def _write_row(self, ws, row_idx: int, row_data: List, is_overdue: bool = False):
        """写入数据行"""
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = self.NORMAL_FONT
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

            if is_overdue:
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    def _write_processes_sheet(self, ws, orders: List[Order], now: datetime, overdue_threshold):
        """写入工序明细sheet"""
        headers = [
            "订单ID", "工序ID", "工序名称", "序号", "状态",
            "接收时间", "发出时间", "是否超期"
        ]
        self._write_header(ws, headers)

        row_idx = 2
        for order in orders:
            records = (
                self.db.query(ProcessRecord)
                .filter(ProcessRecord.order_id == order.order_id)
                .all()
            )

            for record in records:
                is_overdue = False
                if record.last_receive_time and not record.last_ship_time:
                    if now - record.last_receive_time > overdue_threshold:
                        is_overdue = True

                row_data = [
                    order.order_id,
                    record.record_id,
                    "",  # process_name - 需要join获取
                    0,   # process_order - 需要join获取
                    record.record_status,
                    record.last_receive_time.strftime("%Y-%m-%d %H:%M:%S") if record.last_receive_time else "",
                    record.last_ship_time.strftime("%Y-%m-%d %H:%M:%S") if record.last_ship_time else "",
                    "是" if is_overdue else "否",
                ]
                self._write_row(ws, row_idx, row_data, is_overdue=is_overdue)
                row_idx += 1

    def _adjust_column_width(self, ws, headers: List[str]):
        """调整列宽"""
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)

    def _get_status_text(self, status) -> str:
        """获取订单状态中文"""
        status_map = {
            "pending": "待处理",
            "in_progress": "进行中",
            "completed": "已完成",
            "cancelled": "已取消",
        }
        status_val = status.value if hasattr(status, 'value') else str(status)
        return status_map.get(status_val, status_val)

    def _check_order_has_overdue(self, order_id: str, now: datetime, overdue_threshold) -> bool:
        """检查订单是否有超期工序"""
        records = self.db.query(ProcessRecord).filter(
            ProcessRecord.order_id == order_id,
            ProcessRecord.last_receive_time.isnot(None),
            ProcessRecord.last_ship_time.is_(None),
        ).all()

        for record in records:
            if now - record.last_receive_time > overdue_threshold:
                return True
        return False